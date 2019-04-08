import bs4
from requests import get
from urllib.request import urlopen as ureq
from urllib.request import urlretrieve as uret
from bs4 import BeautifulSoup as soup
from openpyxl import *
from tkinter import *
import time
import re
from selenium import webdriver
import os
import datetime




#script for GUI
root = Tk()

root.title('EPL Player Stat Importer')

urlframe = Frame()
urlframe.pack(side = TOP)

def shooting():
    myurl = urlentry.get()

    uclient = ureq(myurl)
    pagehtml = uclient.read()

    # parsing as html
    pagesoup = soup(pagehtml, "html.parser")

    workbook = 'shootingtest.xlsx'

    playerdb = load_workbook(workbook)

    sheet = playerdb.get_sheet_by_name('Data')

    eventdetails = pagesoup.find("div", {"class", "whitebg"})
    event = eventdetails.find("p", {"class": "row3"}).text
    stage = eventdetails.find("p", {"class": "row4"}).text
    datetime = eventdetails.find("p", {"class": "row5"}).text
    datetime = re.split(',', datetime)
    date = str(datetime[0])
    timeplayed = datetime[1]

    print(date)

    timeplayed = re.findall(r'[0-9]{2}:[0-9]{2}', timeplayed)
    print(timeplayed)

    print('Event: ' + event + ' ' + stage)
    if stage == 'Qualification':
        print('Qualification Stages')

        for x in range(1, 38):
            global maxrow
            maxrow = sheet.max_row
            print('Writing to row ' + str(maxrow) + '...')

            test = pagesoup.find("tbody", {"rank": x})
            name = test.find("td", {"colspan": "4"}).text

            base = pagesoup.find("tbody", {"rank": x}).text
            base = base.split()
            print(base)
            country = re.findall(r'[A-Z]{3}', base[-1])

            series = re.sub(r'[A-Za-z]', "", base[-1])
            series = re.findall(r'[0-9]+.[0-9]{1}', series)

            if not country:
                print('no country found')
                country = re.findall(r'[A-Z]{3}', base[-2])

                series = re.sub(r'[A-Za-z]', "", base[-2])
                series = re.findall(r'[0-9]+.[0-9]{1}', series)

            print(country)
            print(series)

            country = country[0]
            print(country)

            score = test.find("td", {"class": "totalcontent"}).text

            eventdetails = pagesoup.find("div", {"class", "whitebg"})
            event = eventdetails.find("p", {"class": "row3"}).text
            stage = eventdetails.find("p", {"class": "row4"}).text

            # test2 = test2.split()

            print(name, score, country, event, stage)

            sheet.cell(row=maxrow + 1, column=1).value = name
            sheet.cell(row=maxrow + 1, column=2).value = score
            sheet.cell(row=maxrow + 1, column=3).value = country

            y = x + 1
            y = str(y)
            sheet.cell(row=maxrow + 1, column=4).value = '=VLOOKUP(C' + y + ',Sheet3!$B$2:$C$247,2,TRUE)'

            for series, count in zip(series, range(1, len(series) + 1)):
                sheet.cell(row=maxrow + 1, column=count + 4).value = series

            time.sleep(2.5)

            playerdb.save(workbook)

            print('/////////////////')

    if stage == 'Final':
        print('Final Stages')

        # this gives us rank id and total score
        test = pagesoup.find_all("td", {"align": "right"})
        print(test)

        # for x in range(1, 38):
        #     maxrow = sheet.max_row
        #     print('Writing to row ' + str(maxrow) + '...')

    # myurl = urlentry.get()
    # uclient = ureq(myurl)
    # pagehtml = uclient.read()
    # # opens browser and scrapes
    # # browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
    # # browser.get(myurl)
    # # time.sleep(3)
    # # pagehtml = browser.page_source
    #
    # # parsing as html
    # pagesoup = soup(pagehtml, "html.parser")
    #
    # # uclient.close()
    # # player details
    # details = pagesoup.find("div", {"class": "personalLists"}).text
    # details = details.split()
    #
    # nationality = details[1]
    # age = details[3]
    # DOB = details[7]
    # height = details[9]
    # weight = details[11]
    # print('Nationality: ' + nationality + '\n' +
    #       'Age: ' + age + '\n' +
    #       'DOB: ' + DOB + '\n' +
    #       'Height: ' + height + '\n' +
    #       'Weight: ' + weight)
    #

def appearance():

    workbook = 'testbook1.xlsx'

    playerdb = load_workbook(workbook)

    sheet = playerdb.get_sheet_by_name('Sheet1')

    textfile = open("linklist.txt")
    print('File Opened')
    lines = textfile.read().split("\n")

    linkslist = []
    for lines in lines:
        if lines[:3] != 'htt':
            continue

        linkslist.append(lines)

    print(linkslist)

    #loop through all link
    for links in linkslist:

        print(str(linkslist.index(links)) + ' out of ' + str(len(linkslist)))

        myurl = links

        uclient = ureq(myurl)
        time.sleep(1.5)

        pagehtml = uclient.read()

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # get match details
        matchtype = pagesoup.find("div", {"class", "details clearfix"}).text
        matchtype = matchtype.split('\n')
        x = matchtype.index('Competition')
        y = matchtype.index('Date')
        competition = matchtype[x + 1]
        date = matchtype[y + 1]

        print(competition)
        print(date)

        # DETERMINE TEAMS
        teams = pagesoup.find_all("h3", {"class": "thick"})
        teamlist = []

        for div in teams:
            x = div.text
            x = ''.join(x.split())
            teamlist.append(x)

        hometeam = teamlist[0]
        awayteam = teamlist[2]

        print('Home Team: ' + hometeam)
        print('Away Team: ' + awayteam)

        if hometeam == 'England':
            print('England Home')
            # MATCH DETAILS
            bigcolumn = pagesoup.find("div", {"class", "combined-lineups-container"})
            leftcolumn = bigcolumn.find("div", {"class", "container left"}).text

            starting = re.findall(r'[A-Z]. [A-Za-z]*', leftcolumn)

            starting.pop()

        if awayteam == 'England':
            print('England Away')
            bigcolumn = pagesoup.find("div", {"class", "combined-lineups-container"})
            leftcolumn = bigcolumn.find("div", {"class", "container right"}).text

            starting = re.findall(r'[A-Z]. [A-Za-z]*', leftcolumn)

            starting.pop()

        print(starting)

        # SUBSTITUTES

        allsubs = bigcolumn.find_next("div", {"class", "combined-lineups-container"})
        if hometeam == 'England':
            actualsubs = allsubs.find("div", {"class", "container left"})
            actualsubs = actualsubs.find("table", {"class", "playerstats lineups substitutions table"}).text

        if awayteam == 'England':
            actualsubs = allsubs.find("div", {"class", "container right"})
            actualsubs = actualsubs.find("table", {"class", "playerstats lineups substitutions table"}).text

        dnpsubs = re.findall(r'[A-Z][.]\s[A-Za-z-]+', actualsubs)
        subsmade = re.findall(r'[A-Z].\s[A-Za-z]+\s+[for]*\s[A-Z][.]\s[A-Za-z-]+\s[0-9]+', actualsubs)
        subs = []

        print(dnpsubs)

        for lines in subsmade:
            sub = re.sub("\n", "", lines)
            sub = re.sub(r"for ", ' for ', sub)
            subs.append(sub)

        print(subs)

        matchdetails = pagesoup.find("div", {"class", "details clearfix"})
        matchdetails2 = matchdetails.find_next("div", {"class", "details clearfix"}).text
        extratimecheck = re.findall(r'\bExtra-time\b', matchdetails2)
        print(extratimecheck)

        if len(extratimecheck) > 0:
            extratime = '1'
            print('Total Playtime = 120 mins (Extra Time)')
        if len(extratimecheck) == 0:
            extratime = '0'
            print('Total Playtime = 90 mins (Normal Time)')

        # we eventually want a dictionary
        playtime = {}

        for subs in subs:
            global timein
            global timeout
            both = re.findall(r'[A-Z][.] [A-Za-z]+', subs)
            timeplayed = re.findall(r'[0-9]+', subs)
            subin = both[0]
            subout = both[1]

            if subin in dnpsubs:
                number = dnpsubs.index(subin)
                dnpsubs.pop(number)

            if subout in dnpsubs:
                number = dnpsubs.index(subout)
                dnpsubs.pop(number)

            if extratime == '1':
                timein = 120 - int(timeplayed[0])
                timeout = timeplayed[0]

            if extratime == '0':
                timein = 90 - int(timeplayed[0])
                timeout = timeplayed[0]

            print(subin, timein)
            print(subout, timeout)

            if subout in starting:
                number = starting.index(subout)
                starting.pop(number)

            playtime[subin] = timein
            playtime[subout] = timeout


        for name in starting:
            if extratime == '1':
                timeplayed = '120'
            if extratime == '0':
                timeplayed = '90'
            print(name, timeplayed)
            playtime[name] = timeplayed

        for name in dnpsubs:
            timeplayed = '0'
            print(name, timeplayed)
            playtime[name] = timeplayed

        print(playtime)

        for key, val in playtime.items():
            global maxrow
            maxrow = sheet.max_row
            sheet.cell(row = maxrow+1, column = 1).value = key
            sheet.cell(row = maxrow+1, column = 2).value = val
            sheet.cell(row=maxrow+1, column=3).value = competition
            sheet.cell(row=maxrow+1, column=4).value = date
            sheet.cell(row = maxrow+1, column=5).value = links

        playerdb.save(workbook)


def test():

    workbook = 'testbook1.xlsx'

    playerdb = load_workbook(workbook)

    sheet = playerdb.get_sheet_by_name('Injury_Off')

    os.system('say "LETS GET IT ON"')

    #get links
    textfile = open("linklist.txt")
    print('File Opened')
    lines = textfile.read().split("\n")

    #List of Daylight Saving dates
    ds2018 = datetime.datetime.strptime('Mar 25, 2018', "%b %d, %Y")
    ds2017 = datetime.datetime.strptime('Mar 26, 2017', "%b %d, %Y")
    ds2016 = datetime.datetime.strptime('Mar 27, 2016', "%b %d, %Y")
    ds2015 = datetime.datetime.strptime('Mar 29, 2015', "%b %d, %Y")
    ds2014 = datetime.datetime.strptime('Mar 30, 2014', "%b %d, %Y")
    ds2013 = datetime.datetime.strptime('Mar 31, 2013', "%b %d, %Y")
    ds2012 = datetime.datetime.strptime('Mar 25, 2012', "%b %d, %Y")
    ds2011 = datetime.datetime.strptime('Mar 27, 2011', "%b %d, %Y")
    ds2010 = datetime.datetime.strptime('Mar 28, 2010', "%b %d, %Y")

    doff2018 = datetime.datetime.strptime('Oct 28, 2018', "%b %d, %Y")
    doff2017 = datetime.datetime.strptime('Oct 29, 2017', "%b %d, %Y")
    doff2016 = datetime.datetime.strptime('Oct 30, 2016', "%b %d, %Y")
    doff2015 = datetime.datetime.strptime('Oct 25, 2015', "%b %d, %Y")
    doff2014 = datetime.datetime.strptime('Oct 26, 2014', "%b %d, %Y")
    doff2013 = datetime.datetime.strptime('Oct 27, 2013', "%b %d, %Y")
    doff2012 = datetime.datetime.strptime('Oct 28, 2012', "%b %d, %Y")
    doff2011 = datetime.datetime.strptime('Oct 30, 2011', "%b %d, %Y")
    doff2010 = datetime.datetime.strptime('Oct 31, 2010', "%b %d, %Y")

    linkslist = []
    for lines in lines:
        if lines[:3] != 'htt':
            continue

        linkslist.append(lines)

    print(linkslist)

    #loop through all link
    for links in linkslist:

        starttime = time.time()

        print(str(linkslist.index(links)) + ' out of ' + str(len(linkslist)))

        global maxrow
        maxrow = sheet.max_row

        noreport = 0
        datecat = 1

        myurl = links

        browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
        browser.get(myurl)
        time.sleep(0.50)
        pagehtml = browser.page_source

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # FOR TESTING
        tabs = pagesoup.find_all("div", {"class", "subnavi_box"})
        statistics = re.findall(r'Statistics..\n.*Statistics', str(tabs))
        statistics = re.findall(r'/.*[0-9]', str(statistics))

        statspagelink = 'https://www.transfermarkt.co.uk' + str(statistics[0])

        print(statspagelink)

        #MATCH DETAILS
        gamescore = pagesoup.find("div", {"class", "sb-endstand"}).text
        gamescore = re.findall(r'[0-9{1,}]:[0-9{1,}]', gamescore)
        gamescore = gamescore[0]
        gamescore = gamescore.split(":")
        homescore = gamescore[0]
        awayscore = gamescore[1]

        totalconceded = int(homescore) + int(awayscore)

        date = pagesoup.find("div", {"class", "sb-spieldaten"})
        date = date.find("p", {"class", "sb-datum hide-for-small"})
        date = re.sub("<.*?>", "", str(date))
        dateplayed = re.findall(r'[a-zA-Z]{3} [0-9]{1,}, [0-9]{4}', date)
        dateplayed = str(dateplayed[0])
        # print(dateplayed)
        dateplayed = datetime.datetime.strptime(dateplayed, "%b %d, %Y")
        # print(dateplayed.year)


        # For calculating days from Daylight Savings Day
        def calculatedate():
            global category
            if datecat != 0:
                if 8 <= int(dsONdiff[0]) <= 14:
                    print('Category A - Pre')
                    category = '0'

                if -8 < int(dsONdiff[0]) < 8:
                    print('Category B - Week of ')
                    category = '1'

                if -14 <= int(dsONdiff[0]) <= -8:
                    print('Category C - Post')
                    category = '2'

                if int(dsONdiff[0]) > 14:
                    print('Category D - DNC')
                    category = '3'

                if int(dsONdiff[0]) < -14:
                    print('Category D - DNC')
                    category = '3'

            if datecat == 0:
                category = '1'


# Calculating Offset data
        if int(dateplayed.year) == int(doff2010.year):
            print(doff2010 - dateplayed)

            if doff2010 != dateplayed:
                dsONdiff = doff2010 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2010 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()

        if int(dateplayed.year) == int(doff2011.year):
            print(doff2011 - dateplayed)

            if doff2011 != dateplayed:
                dsONdiff = doff2011 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2011 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2012.year):
            if doff2012 != dateplayed:
                dsONdiff = doff2012 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2012 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2013.year):
            if doff2013 != dateplayed:
                dsONdiff = doff2013 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2013 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2014.year):
            if doff2014 != dateplayed:
                dsONdiff = doff2014 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2014 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2015.year):
            if doff2015 != dateplayed:
                dsONdiff = doff2015 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2015 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2016.year):
            if doff2016 != dateplayed:
                dsONdiff = doff2016 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2016 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


        if int(dateplayed.year) == int(doff2017.year):
            if doff2017 != dateplayed:
                dsONdiff = doff2017 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2017 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()

        if int(dateplayed.year) == int(doff2018.year):
            if doff2018 != dateplayed:
                dsONdiff = doff2018 - dateplayed

                dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
                dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
                print(dsONdiff)

                calculatedate()

            if doff2018 == dateplayed:
                dsONdiff = 0
                datecat = 0
                calculatedate()


# EXTRACTING DLS ONSET DATA
#         if int(dateplayed.year) == int(ds2010.year):
#             print(ds2010 - dateplayed)
#
#             if ds2010 != dateplayed:
#                 dsONdiff = ds2010 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2010 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#         if int(dateplayed.year) == int(ds2011.year):
#             print(ds2011 - dateplayed)
#
#             if ds2011 != dateplayed:
#                 dsONdiff = ds2011 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2011 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2012.year):
#             if ds2012 != dateplayed:
#                 dsONdiff = ds2012 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2012 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2013.year):
#             if ds2013 != dateplayed:
#                 dsONdiff = ds2013 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2013 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2014.year):
#             if ds2014 != dateplayed:
#                 dsONdiff = ds2014 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2014 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2015.year):
#             if ds2015 != dateplayed:
#                 dsONdiff = ds2015 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2015 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2016.year):
#             if ds2016 != dateplayed:
#                 dsONdiff = ds2016 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2016 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#
#         if int(dateplayed.year) == int(ds2017.year):
#             if ds2017 != dateplayed:
#                 dsONdiff = ds2017 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2017 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()
#
#         if int(dateplayed.year) == int(ds2018.year):
#             if ds2018 != dateplayed:
#                 dsONdiff = ds2018 - dateplayed
#
#                 dsONdiff = re.findall(r'[-0-9]{1,} day', str(dsONdiff))
#                 dsONdiff = re.findall(r'[-0-9]{1,}', str(dsONdiff))
#                 print(dsONdiff)
#
#                 calculatedate()
#
#             if ds2018 == dateplayed:
#                 dsONdiff = 0
#                 datecat = 0
#                 calculatedate()


        #Extract time that match was played
        timeplayed = re.findall(r'[0-9]+:[0-9]{2}', date)
        timeplayed = timeplayed[0] + ' PM'

        league = pagesoup.find("div", {"class", "spielername-profil"}).text
        league = re.sub("\n", "", league)
        league = str(league)
        print(league)

        #INJURY DETAILS
        matchdetails = pagesoup.find_all("div", {"class", "sb-ereignisse"})
        breakdown = re.sub("<.*?>", "", str(matchdetails))

        #for fouls
        yellowcards = re.findall('Yellow card', breakdown)
        redcards = re.findall('Red card', breakdown)

        yellowcards = str(len(yellowcards))
        redcards = str(len(redcards))


        #for injuries
        injuries = re.findall('Injury', breakdown)
        injuries = str(len(injuries))
        print(injuries)

        #not reported findings
        notreported = re.findall('Not reported', breakdown)
        if len(notreported) > 0:
            noreport = 1

        print('Date Played: ' + str(dateplayed))
        print('Time Played: ' + timeplayed)
        print('Amount of Injuries: ' + injuries)
        print('Yellow Cards: ' + yellowcards)
        print('Red Cards: ' + redcards)
        print('Home Team Score: ' + homescore)
        print('Away Team Score: ' + awayscore)
        print('Total Goals Conceded: ' + str(totalconceded))


        ## This portion writes to excel, comment out when testing

        sheet.cell(row=maxrow + 1, column=1).value = dateplayed
        sheet.cell(row=maxrow + 1, column=2).value = timeplayed
        sheet.cell(row=maxrow + 1, column=3).value = injuries
        sheet.cell(row=maxrow + 1, column=4).value = league
        sheet.cell(row=maxrow + 1, column=5).value = yellowcards
        sheet.cell(row=maxrow + 1, column=6).value = redcards
        sheet.cell(row=maxrow + 1, column=7).value = links
        sheet.cell(row=maxrow + 1, column=9).value = homescore
        sheet.cell(row=maxrow + 1, column=10).value = awayscore
        sheet.cell(row=maxrow + 1, column=11).value = str(totalconceded)
        sheet.cell(row=maxrow + 1, column=12).value = category

        if datecat != 0:
            sheet.cell(row=maxrow + 1, column=13).value = str(dsONdiff[0])

        if datecat == 0:
            sheet.cell(row=maxrow + 1, column=13).value = '0'
            print('DAY OF DLS')





        if noreport == 1:
            sheet.cell(row=maxrow + 1, column=8).value = '1'
            print('Missing Data: YES')

        if noreport != 1:
            sheet.cell(row=maxrow + 1, column=8).value = '0'
            print('Missing Data: NO')


        noreport = 0
        datecat = 1

        browser.close()

        # Open new browser to get match statistics
        browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
        myurl = statspagelink
        browser.get(myurl)
        time.sleep(0.50)
        pagehtml = browser.page_source

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # Get match statistics
        matchstatistics = pagesoup.find_all("div", {"class", "sb-statistik"})
        matchstatistics = re.findall(r'>[0-9]{1,}<', str(matchstatistics))
        matchstatsclean = []
        for i in matchstatistics:
            i = re.sub("[><]", "", i)
            print(i)
            matchstatsclean.append(i)
        print(matchstatsclean)
        print(len(matchstatistics))

        counting = 0
        missingdata = matchstatsclean.count('0')

        for i in matchstatsclean:
            if len(matchstatsclean) == 0:
                break
            if missingdata == 14:
                break

            counting += 1
            sheet.cell(row = maxrow + 1, column = 13+counting).value = i

        browser.close()
        playerdb.save(workbook)


        endtime = time.time()
        onelooptime = round(endtime-starttime,2)



        print('This loop took: ' + str(onelooptime) + ' secs.')

        print('//////////////////////////////////////////////')




def multiscrape():
    def findlinks():
        textfile = open("linklist.txt")
        print('File Opened')
        lines = textfile.read().split("\n")
        global linkslist
        global linkslist2
        linkslist = []
        linkslist2 = []
        for lines in lines:
            if lines[:3] != 'htt':
                continue
            # one list for overview (height, nationality, DOB, age stats)
            linkslist2.append(lines)

            #this list is is for stats page
            lines = lines[:-9]
            # if 17/18 data wanted, edit url
            lines = lines + 'stats?co=1&se=21'
            linkslist.append(lines)

        print(linkslist)
        print(linkslist2)

    workbook = 'EPL 1213.xlsx'

    findlinks()
    playerdb = load_workbook(workbook)

    # for links, count in zip(linkslist2, range(1,len(linkslist2)+1)):
    #     myurl = links
    #     uclient = ureq(myurl)
    #     pagehtml = uclient.read()
    #     # opens browser and scrapes
    #     # browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
    #     # browser.get(myurl)
    #     # time.sleep(3)
    #     # pagehtml = browser.page_source
    #
    #     # parsing as html
    #     pagesoup = soup(pagehtml, "html.parser")
    #
    #     # uclient.close()
    #     # player details
    #     details = pagesoup.find("div", {"class": "personalLists"}).text
    #     details = details.split()
    #     if len(details) < 7:
    #         nationality = details[1]
    #         age = '0'
    #         DOB = details[5]
    #         YOB = details[5][6:]
    #         height = '0'
    #         weight = '0'
    #     if len(details) > 7:
    #         nationality = details[1]
    #         age = details[3]
    #         DOB = details[7]
    #         YOB = details[7][6:]
    #         height = details[9]
    #         weight = details[11]
    #     print('Nationality: ' + nationality + '\n' +
    #           'Age: ' + age + '\n' +
    #           'DOB: ' + DOB + '\n' +
    #           'Year Birth: ' + YOB + '\n' +
    #           'Height: ' + height + '\n' +
    #           'Weight: ' + weight + '\n' +
    #           '///////////////////////////////////')
    #
    #
    #     time.sleep(1.5)

    for links, count in zip(linkslist, range(1,len(linkslist)+1)):
        #for personal details
        myurl1 = linkslist2[count-1]
        uclient = ureq(myurl1)
        pagehtml = uclient.read()
        # opens browser and scrapes
        # browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
        # browser.get(myurl)
        # time.sleep(3)
        # pagehtml = browser.page_source

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # uclient.close()
        # player details
        details = pagesoup.find("div", {"class": "personalLists"}).text
        details = details.split()
        if len(details) < 7:
            nationality = details[1]
            age = '0'
            DOB = details[5]
            YOB = details[5][6:]
            height = '0'
            weight = '0'
        if len(details) > 7:
            nationality = details[1]
            age = details[3]
            DOB = details[7]
            YOB = details[7][6:]
            height = details[9]
            weight = details[11]
        print('Nationality: ' + nationality + '\n' +
              'Age: ' + age + '\n' +
              'DOB: ' + DOB + '\n' +
              'Year Birth: ' + YOB + '\n' +
              'Height: ' + height + '\n' +
              'Weight: ' + weight + '\n' +
              '///////////////////////////////////')

        time.sleep(1.5)



        myurl = links
        print('Scrape No. ' + str(count) + ' out of ' + str(len(linkslist)))
        # opens browser and scrapes
        browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
        browser.get(myurl)
        time.sleep(3.50)
        pagehtml = browser.page_source

        # parsing as html
        pagesoup = soup(pagehtml, "html.parser")

        # uclient.close()
        # player details
        name = pagesoup.find("div", {"class": "name"}).text
        try:
            jerseyno = pagesoup.find("div", {"class": "number"}).text
        except:
            jerseyno = '0'

        positionget = pagesoup.find_all("div", {"class": "info"})

        global position
        position = 'null'
        for div in positionget:
            x = div.text
            x = ''.join(x.split())
            if x == 'Goalkeeper':
                position = 'Goalkeeper'
            if x == 'Defender':
                position = 'Defender'
            if x == 'Midfielder':
                position = 'Midfielder'
            if x == 'Forward':
                position = 'Forward'

            print('Position: ' + position)

        team = pagesoup.find("div", {"class": "info"}).text
        if team == position:
            team = '0'

        if position == 'Defender':
            # determine sheet
            sheet = playerdb.get_sheet_by_name('Defender')

            # determine number of rows existing, write to that row number plus one
            global maxrow
            maxrow = sheet.max_row
            print('Number of rows is ' + str(maxrow))

            # general stats
            appearances = pagesoup.find("span", {"class": "allStatContainer statappearances"}).text
            wins = pagesoup.find("span", {"class": "allStatContainer statwins"}).text
            losses = pagesoup.find("span", {"class": "allStatContainer statlosses"}).text

            # defence stats
            try:
                cleansheet = pagesoup.find("span", {"class": "allStatContainer statclean_sheet"}).text
            except:
                cleansheet = '0'
            try:
                goalconceded = pagesoup.find("span", {"class": "allStatContainer statgoals_conceded"}).text
            except:
                goalconceded = '0'
            tackles = pagesoup.find("span", {"class": "allStatContainer stattotal_tackle"}).text
            tacklessuccess = pagesoup.find("span", {"class": "allStatContainer stattackle_success"}).text
            try:
                lastmantackle = pagesoup.find("span", {"class": "allStatContainer statlast_man_tackle"}).text
            except:
                lastmantackle = '0'

            blocks = pagesoup.find("span", {"class": "allStatContainer statblocked_scoring_att"}).text
            interceptions = pagesoup.find("span", {"class": "allStatContainer statinterception"}).text
            clearances = pagesoup.find("span", {"class": "allStatContainer stattotal_clearance"}).text
            headedclearance = pagesoup.find("span", {"class": "allStatContainer stateffective_head_clearance"}).text
            try:
                clearanceoffline = pagesoup.find("span", {"class": "allStatContainer statclearance_off_line"}).text
            except:
                clearanceoffline = '0'
            recovery = pagesoup.find("span", {"class": "allStatContainer statball_recovery"}).text
            duelswon = pagesoup.find("span", {"class": "allStatContainer statduel_won"}).text
            duelslost = pagesoup.find("span", {"class": "allStatContainer statduel_lost"}).text
            fiftyfiftywon = pagesoup.find("span", {"class": "allStatContainer statwon_contest"}).text
            aerialwon = pagesoup.find("span", {"class": "allStatContainer stataerial_won"}).text
            aeriallost = pagesoup.find("span", {"class": "allStatContainer stataerial_lost"}).text
            try:
                owngoals = pagesoup.find("span", {"class": "allStatContainer statown_goals"}).text
            except:
                owngoals = '0'
            errortogoal = pagesoup.find("span", {"class": "allStatContainer staterror_lead_to_goal"}).text

            # discipline stats
            yellowcard = pagesoup.find("span", {"class": "allStatContainer statyellow_card"}).text
            redcard = pagesoup.find("span", {"class": "allStatContainer statred_card"}).text
            fouls = pagesoup.find("span", {"class": "allStatContainer statfouls"}).text
            offsides = pagesoup.find("span", {"class": "allStatContainer stattotal_offside"}).text

            # teamplay stats
            assists = pagesoup.find("span", {"class": "allStatContainer statgoal_assist"}).text
            passes = pagesoup.find("span", {"class": "allStatContainer stattotal_pass"}).text
            passespergame = pagesoup.find("span", {"class": "allStatContainer stattotal_pass_per_game"}).text
            bigchancecreated = pagesoup.find("span", {"class": "allStatContainer statbig_chance_created"}).text
            crosses = pagesoup.find("span", {"class": "allStatContainer stattotal_cross"}).text
            crossacc = pagesoup.find("span", {"class": "allStatContainer statcross_accuracy"}).text
            throughballs = pagesoup.find("span", {"class": "allStatContainer stattotal_through_ball"}).text
            acclongballs = pagesoup.find("span", {"class": "allStatContainer stataccurate_long_balls"}).text

            # attack stats
            goals = pagesoup.find("span", {"class": "allStatContainer statgoals"}).text
            headgoals = pagesoup.find("span", {"class": "allStatContainer statatt_hd_goal"}).text
            goalsright = pagesoup.find("span", {"class": "allStatContainer statatt_rf_goal"}).text
            goalsleft = pagesoup.find("span", {"class": "allStatContainer statatt_lf_goal"}).text
            woodwork = pagesoup.find("span", {"class": "allStatContainer stathit_woodwork"}).text

            # using lists to trim text and for writing to CSV later
            playerdetails = [name, jerseyno, position, team]
            generallist = [appearances, wins, losses]
            defendinglist = [cleansheet, goalconceded, tackles, tacklessuccess, lastmantackle, blocks, interceptions,
                             clearances, headedclearance, clearanceoffline, recovery, duelswon, duelslost,
                             fiftyfiftywon, aerialwon, aeriallost, owngoals, errortogoal]
            disciplinelist = [yellowcard, redcard, fouls, offsides]
            teamplaylist = [assists, passes, passespergame, bigchancecreated, crosses, crossacc, throughballs,
                            acclongballs]
            attacklist = [goals, headgoals, goalsright, goalsleft, woodwork]
            detailslist = [nationality, age, DOB, YOB, height, weight]

            for x, y in zip(playerdetails, range(0, len(playerdetails))):
                if x != name:
                    x = ''.join(x.split())
                    x = x.replace(',', '')
                    playerdetails[y] = x

                if x == name:
                    print('Player: ' + name + ' ' + position)
                    playerdetails[y] = x

            for x, y in zip(generallist, range(0, len(generallist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                generallist[y] = x

            for x, y in zip(defendinglist, range(0, len(defendinglist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                defendinglist[y] = x

            for x, y in zip(attacklist, range(0, len(attacklist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                attacklist[y] = x

            for x, y in zip(disciplinelist, range(0, len(disciplinelist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                disciplinelist[y] = x

            for x, y in zip(teamplaylist, range(0, len(teamplaylist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                teamplaylist[y] = x

            print(playerdetails)
            print(generallist)
            print(defendinglist)
            print(attacklist)
            print(disciplinelist)
            print(teamplaylist)
            print(detailslist)
            # to remove all whitespace
            # saves = ''.join(saves.split())

            # write to the csv file
            for x, y in zip(playerdetails, range(1, len(playerdetails) + 1)):
                sheet.cell(row=maxrow + 1, column=y).value = x

            for x, y in zip(generallist, range(1, len(generallist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 4).value = x

            for x, y in zip(defendinglist, range(1, len(defendinglist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 7).value = x

            for x, y in zip(attacklist, range(1, len(attacklist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 25).value = x

            for x, y in zip(disciplinelist, range(1, len(disciplinelist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 30).value = x

            for x, y in zip(teamplaylist, range(1, len(teamplaylist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 34).value = x

            for x, y in zip(detailslist, range(1, len(detailslist)+ 1)):
                sheet.cell(row = maxrow + 1, column = y + 42).value = x

            playerdb.save(workbook)

        if position == 'Goalkeeper':
            # determine sheet
            sheet = playerdb.get_sheet_by_name('Goalkeeper')

            # determine number of rows existing, write to that row number plus one
            maxrow = sheet.max_row
            print('number of rows is ' + str(maxrow))

            # general stats
            appearances = pagesoup.find("span", {"class": "allStatContainer statappearances"}).text
            wins = pagesoup.find("span", {"class": "allStatContainer statwins"}).text
            losses = pagesoup.find("span", {"class": "allStatContainer statlosses"}).text

            # goalkeeping stats
            saves = pagesoup.find("span", {"class": "allStatContainer statsaves"}).text
            penaltysaves = pagesoup.find("span", {"class": "allStatContainer statpenalty_save"}).text
            punches = pagesoup.find("span", {"class": "allStatContainer statpunches"}).text
            highclaim = pagesoup.find("span", {"class": "allStatContainer statgood_high_claim"}).text
            catches = pagesoup.find("span", {"class": "allStatContainer statcatches"}).text
            sweepclearance = pagesoup.find("span", {"class": "allStatContainer stattotal_keeper_sweeper"}).text
            throws = pagesoup.find("span", {"class": "allStatContainer statkeeper_throws"}).text
            goalkicks = pagesoup.find("span", {"class": "allStatContainer statgoal_kicks"}).text

            # defence stats
            cleansheet = pagesoup.find("span", {"class": "allStatContainer statclean_sheet"}).text
            goalconceded = pagesoup.find("span", {"class": "allStatContainer statgoals_conceded"}).text
            errortogoal = pagesoup.find("span", {"class": "allStatContainer staterror_lead_to_goal"}).text
            owngoal = pagesoup.find("span", {"class": "allStatContainer statown_goals"}).text

            # discipline stats
            yellowcard = pagesoup.find("span", {"class": "allStatContainer statyellow_card"}).text
            redcard = pagesoup.find("span", {"class": "allStatContainer statred_card"}).text
            fouls = pagesoup.find("span", {"class": "allStatContainer statfouls"}).text

            # teamplay stats
            goals = pagesoup.find("span", {"class": "allStatContainer statgoals"}).text
            assists = pagesoup.find("span", {"class": "allStatContainer statgoal_assist"}).text
            passes = pagesoup.find("span", {"class": "allStatContainer stattotal_pass"}).text
            passespergame = pagesoup.find("span", {"class": "allStatContainer stattotal_pass_per_game"}).text
            longballs = pagesoup.find("span", {"class": "allStatContainer stataccurate_long_balls"}).text

            # using lists to trim text and for writing to CSV later
            playerdetails = [name, jerseyno, position, team]
            generallist = [appearances, wins, losses]
            goalkeepinglist = [saves, penaltysaves, punches, highclaim, catches, sweepclearance, throws, goalkicks]
            defencelist = [cleansheet, goalconceded, errortogoal, owngoal]
            disciplinelist = [yellowcard, redcard, fouls]
            teamplaylist = [goals, assists, passes, passespergame, longballs]
            detailslist = [nationality, age, DOB, YOB, height, weight]


            for x, y in zip(playerdetails, range(0, len(playerdetails))):
                if x != name:
                    x = ''.join(x.split())
                    x = x.replace(',', '')
                    playerdetails[y] = x

                if x == name:
                    print('Player: ' + name + ' ' + position)
                    playerdetails[y] = x

            for x, y in zip(generallist, range(0, len(generallist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                generallist[y] = x

            for x, y in zip(goalkeepinglist, range(0, len(goalkeepinglist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                goalkeepinglist[y] = x

            for x, y in zip(defencelist, range(0, len(defencelist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                defencelist[y] = x

            for x, y in zip(disciplinelist, range(0, len(disciplinelist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                disciplinelist[y] = x

            for x, y in zip(teamplaylist, range(0, len(teamplaylist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                teamplaylist[y] = x

            print(playerdetails)
            print(generallist)
            print(goalkeepinglist)
            print(defencelist)
            print(disciplinelist)
            print(teamplaylist)

            # to remove all whitespace
            # saves = ''.join(saves.split())

            # write to the csv file
            for x, y in zip(playerdetails, range(1, len(playerdetails) + 1)):
                sheet.cell(row=maxrow + 1, column=y).value = x

            for x, y in zip(generallist, range(1, len(generallist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 4).value = x

            for x, y in zip(goalkeepinglist, range(1, len(goalkeepinglist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 7).value = x

            for x, y in zip(defencelist, range(1, len(defencelist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 15).value = x

            for x, y in zip(disciplinelist, range(1, len(disciplinelist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 19).value = x

            for x, y in zip(teamplaylist, range(1, len(teamplaylist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 22).value = x

            for x, y in zip(detailslist, range(1, len(detailslist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 27).value = x
            playerdb.save(workbook)

        if position == 'Midfielder':
            # determine sheet
            sheet = playerdb.get_sheet_by_name('Midfielder')

            # determine number of rows existing, write to that row number plus one
            maxrow = sheet.max_row
            print('number of rows is ' + str(maxrow))

            # general stats
            appearances = pagesoup.find("span", {"class": "allStatContainer statappearances"}).text
            wins = pagesoup.find("span", {"class": "allStatContainer statwins"}).text
            losses = pagesoup.find("span", {"class": "allStatContainer statlosses"}).text

            # defence stats
            tackles = pagesoup.find("span", {"class": "allStatContainer stattotal_tackle"}).text
            try:
                tacklessuccess = pagesoup.find("span", {"class": "allStatContainer stattackle_success"}).text
            except:
                tacklessuccess = '0'
            blocks = pagesoup.find("span", {"class": "allStatContainer statblocked_scoring_att"}).text
            interceptions = pagesoup.find("span", {"class": "allStatContainer statinterception"}).text
            clearances = pagesoup.find("span", {"class": "allStatContainer stattotal_clearance"}).text
            headedclearance = pagesoup.find("span", {"class": "allStatContainer stateffective_head_clearance"}).text
            try:
                recovery = pagesoup.find("span", {"class": "allStatContainer statball_recovery"}).text
            except:
                recovery = '0'
            try:
                duelswon = pagesoup.find("span", {"class": "allStatContainer statduel_won"}).text
            except:
                duelswon = '0'
            try:
                duelslost = pagesoup.find("span", {"class": "allStatContainer statduel_lost"}).text
            except:
                duelslost = '0'
            try:
                fiftyfiftywon = pagesoup.find("span", {"class": "allStatContainer statwon_contest"}).text
            except:
                fiftyfiftywon = '0'
            try:
                aerialwon = pagesoup.find("span", {"class": "allStatContainer stataerial_won"}).text
            except:
                aerialwon = '0'
            try:
                aeriallost = pagesoup.find("span", {"class": "allStatContainer stataerial_lost"}).text
            except:
                aeriallost = '0'
            try:
                errortogoal = pagesoup.find("span", {"class": "allStatContainer staterror_lead_to_goal"}).text
            except:
                errortogoal = '0'

            # discipline stats
            yellowcard = pagesoup.find("span", {"class": "allStatContainer statyellow_card"}).text
            redcard = pagesoup.find("span", {"class": "allStatContainer statred_card"}).text
            fouls = pagesoup.find("span", {"class": "allStatContainer statfouls"}).text
            offsides = pagesoup.find("span", {"class": "allStatContainer stattotal_offside"}).text

            # teamplay stats
            assists = pagesoup.find("span", {"class": "allStatContainer statgoal_assist"}).text
            passes = pagesoup.find("span", {"class": "allStatContainer stattotal_pass"}).text
            passespergame = pagesoup.find("span", {"class": "allStatContainer stattotal_pass_per_game"}).text
            bigchancecreated = pagesoup.find("span", {"class": "allStatContainer statbig_chance_created"}).text
            crosses = pagesoup.find("span", {"class": "allStatContainer stattotal_cross"}).text
            try:
                crossacc = pagesoup.find("span", {"class": "allStatContainer statcross_accuracy"}).text
            except:
                crossacc = '0'
            try:
                throughballs = pagesoup.find("span", {"class": "allStatContainer stattotal_through_ball"}).text
            except:
                throughballs = '0'
            try:
                acclongballs = pagesoup.find("span", {"class": "allStatContainer stataccurate_long_balls"}).text
            except:
                acclongballs = '0'

            # attack stats
            goals = pagesoup.find("span", {"class": "allStatContainer statgoals"}).text
            goalspg = pagesoup.find("span", {"class": "allStatContainer statgoals_per_game"}).text
            headgoals = pagesoup.find("span", {"class": "allStatContainer statatt_hd_goal"}).text
            goalsright = pagesoup.find("span", {"class": "allStatContainer statatt_rf_goal"}).text
            goalsleft = pagesoup.find("span", {"class": "allStatContainer statatt_lf_goal"}).text
            penaltyscored = pagesoup.find("span", {"class": "allStatContainer statatt_pen_goal"}).text
            freekickscored = pagesoup.find("span", {"class": "allStatContainer statatt_freekick_goal"}).text
            shots = pagesoup.find("span", {"class": "allStatContainer stattotal_scoring_att"}).text
            shotstarget = pagesoup.find("span", {"class": "allStatContainer statontarget_scoring_att"}).text
            shotacc = pagesoup.find("span", {"class": "allStatContainer statshot_accuracy"}).text
            woodwork = pagesoup.find("span", {"class": "allStatContainer stathit_woodwork"}).text
            bigchancemiss = pagesoup.find("span", {"class": "allStatContainer statbig_chance_missed"}).text

            # using lists to trim text and for writing to CSV later
            playerdetails = [name, jerseyno, position, team]
            generallist = [appearances, wins, losses]
            defendinglist = [tackles, tacklessuccess, blocks, interceptions, clearances, headedclearance, recovery,
                             duelswon, duelslost, fiftyfiftywon, aerialwon, aeriallost, errortogoal]
            disciplinelist = [yellowcard, redcard, fouls, offsides]
            teamplaylist = [assists, passes, passespergame, bigchancecreated, crosses, crossacc, throughballs,
                            acclongballs]
            attacklist = [goals, goalspg, headgoals, goalsright, goalsleft, penaltyscored, freekickscored, shots,
                          shotstarget, shotacc, woodwork, bigchancemiss]
            detailslist = [nationality, age, DOB, YOB, height, weight]

            print(len(defendinglist))
            print(len(attacklist))
            print(len(disciplinelist))
            print(len(teamplaylist))

            for x, y in zip(playerdetails, range(0, len(playerdetails))):
                if x != name:
                    x = ''.join(x.split())
                    x = x.replace(',', '')
                    playerdetails[y] = x

                if x == name:
                    print('Player: ' + name + ' ' + position)
                    playerdetails[y] = x

            for x, y in zip(generallist, range(0, len(generallist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                generallist[y] = x

            for x, y in zip(defendinglist, range(0, len(defendinglist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                defendinglist[y] = x

            for x, y in zip(attacklist, range(0, len(attacklist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                attacklist[y] = x

            for x, y in zip(disciplinelist, range(0, len(disciplinelist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                disciplinelist[y] = x

            for x, y in zip(teamplaylist, range(0, len(teamplaylist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                teamplaylist[y] = x

            print(playerdetails)
            print(generallist)
            print(defendinglist)
            print(attacklist)
            print(disciplinelist)
            print(teamplaylist)

            # to remove all whitespace
            # saves = ''.join(saves.split())

            # write to the csv file
            for x, y in zip(playerdetails, range(1, len(playerdetails) + 1)):
                sheet.cell(row=maxrow + 1, column=y).value = x

            for x, y in zip(generallist, range(1, len(generallist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 4).value = x

            for x, y in zip(defendinglist, range(1, len(defendinglist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 7).value = x

            for x, y in zip(disciplinelist, range(1, len(disciplinelist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 20).value = x

            for x, y in zip(attacklist, range(1, len(attacklist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 24).value = x

            for x, y in zip(teamplaylist, range(1, len(teamplaylist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 36).value = x

            for x, y in zip(detailslist, range(1, len(detailslist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 44).value = x

            playerdb.save(workbook)

        if position == 'Forward':
            # determine sheet
            sheet = playerdb.get_sheet_by_name('Forward')

            # determine number of rows existing, write to that row number plus one
            maxrow = sheet.max_row
            print('number of rows is ' + str(maxrow))

            # general stats
            appearances = pagesoup.find("span", {"class": "allStatContainer statappearances"}).text
            wins = pagesoup.find("span", {"class": "allStatContainer statwins"}).text
            losses = pagesoup.find("span", {"class": "allStatContainer statlosses"}).text

            # defence stats 5
            tackles = pagesoup.find("span", {"class": "allStatContainer stattotal_tackle"}).text
            blocks = pagesoup.find("span", {"class": "allStatContainer statblocked_scoring_att"}).text
            interceptions = pagesoup.find("span", {"class": "allStatContainer statinterception"}).text
            clearances = pagesoup.find("span", {"class": "allStatContainer stattotal_clearance"}).text
            headedclearance = pagesoup.find("span", {"class": "allStatContainer stateffective_head_clearance"}).text

            # discipline stats 4
            yellowcard = pagesoup.find("span", {"class": "allStatContainer statyellow_card"}).text
            redcard = pagesoup.find("span", {"class": "allStatContainer statred_card"}).text
            fouls = pagesoup.find("span", {"class": "allStatContainer statfouls"}).text
            offsides = pagesoup.find("span", {"class": "allStatContainer stattotal_offside"}).text

            # teamplay stats 5
            assists = pagesoup.find("span", {"class": "allStatContainer statgoal_assist"}).text
            passes = pagesoup.find("span", {"class": "allStatContainer stattotal_pass"}).text
            passespergame = pagesoup.find("span", {"class": "allStatContainer stattotal_pass_per_game"}).text
            bigchancecreated = pagesoup.find("span", {"class": "allStatContainer statbig_chance_created"}).text
            crosses = pagesoup.find("span", {"class": "allStatContainer stattotal_cross"}).text

            # attack stats 12
            goals = pagesoup.find("span", {"class": "allStatContainer statgoals"}).text
            goalspg = pagesoup.find("span", {"class": "allStatContainer statgoals_per_game"}).text
            headgoals = pagesoup.find("span", {"class": "allStatContainer statatt_hd_goal"}).text
            goalsright = pagesoup.find("span", {"class": "allStatContainer statatt_rf_goal"}).text
            goalsleft = pagesoup.find("span", {"class": "allStatContainer statatt_lf_goal"}).text
            penaltyscored = pagesoup.find("span", {"class": "allStatContainer statatt_pen_goal"}).text
            freekickscored = pagesoup.find("span", {"class": "allStatContainer statatt_freekick_goal"}).text
            shots = pagesoup.find("span", {"class": "allStatContainer stattotal_scoring_att"}).text
            shotstarget = pagesoup.find("span", {"class": "allStatContainer statontarget_scoring_att"}).text
            shotacc = pagesoup.find("span", {"class": "allStatContainer statshot_accuracy"}).text
            woodwork = pagesoup.find("span", {"class": "allStatContainer stathit_woodwork"}).text
            bigchancemiss = pagesoup.find("span", {"class": "allStatContainer statbig_chance_missed"}).text

            # using lists to trim text and for writing to CSV later
            playerdetails = [name, jerseyno, position, team]
            generallist = [appearances, wins, losses]
            defendinglist = [tackles, blocks, interceptions, clearances, headedclearance]
            disciplinelist = [yellowcard, redcard, fouls, offsides]
            teamplaylist = [assists, passes, passespergame, bigchancecreated, crosses]
            attacklist = [goals, goalspg, headgoals, goalsright, goalsleft, penaltyscored, freekickscored, shots,
                          shotstarget, shotacc, woodwork, bigchancemiss]
            detailslist = [nationality, age, DOB, YOB, height, weight]


            for x, y in zip(playerdetails, range(0, len(playerdetails))):
                if x != name:
                    x = ''.join(x.split())
                    x = x.replace(',', '')
                    playerdetails[y] = x

                if x == name:
                    print('Player: ' + name + ' ' + position)
                    playerdetails[y] = x

            for x, y in zip(generallist, range(0, len(generallist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                generallist[y] = x

            for x, y in zip(defendinglist, range(0, len(defendinglist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                defendinglist[y] = x

            for x, y in zip(attacklist, range(0, len(attacklist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                attacklist[y] = x

            for x, y in zip(disciplinelist, range(0, len(disciplinelist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                disciplinelist[y] = x

            for x, y in zip(teamplaylist, range(0, len(teamplaylist))):
                x = ''.join(x.split())
                x = x.replace(',', '')
                teamplaylist[y] = x

            print(playerdetails)
            print(generallist)
            print(defendinglist)
            print(attacklist)
            print(disciplinelist)
            print(teamplaylist)

            # to remove all whitespace
            # saves = ''.join(saves.split())

            # write to the csv file
            for x, y in zip(playerdetails, range(1, len(playerdetails) + 1)):
                sheet.cell(row=maxrow + 1, column=y).value = x

            for x, y in zip(generallist, range(1, len(generallist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 4).value = x

            for x, y in zip(defendinglist, range(1, len(defendinglist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 7).value = x

            for x, y in zip(disciplinelist, range(1, len(disciplinelist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 12).value = x

            for x, y in zip(attacklist, range(1, len(attacklist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 16).value = x

            for x, y in zip(teamplaylist, range(1, len(teamplaylist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 28).value = x

            for x, y in zip(detailslist, range(1, len(detailslist) + 1)):
                sheet.cell(row=maxrow + 1, column=y + 33).value = x

        playerdb.save(workbook)
        browser.close()

def startscrape():
    playerdb = load_workbook('EPL 1617.xlsx')

    #url to be scraped
    myurl = urlentry.get()

    #adds text to URL for 2017/18 data, delete if want full stats
    # myurl = myurl + '?co=1&se=54'

    #opens browser and scrapes
    browser = webdriver.Chrome("/Users/Qixiang/Dropbox/ICS/venv/chromedriver")
    browser.get(myurl)
    time.sleep(3)
    pagehtml = browser.page_source

    # uret(myurl, 'testpage.html')
    #
    # uclient = ureq(myurl)
    # pagehtml = uclient.read()

    #parsing as html
    pagesoup = soup(pagehtml, "html.parser")

    # uclient.close()
    #player details
    name = pagesoup.find("div",{"class":"name"}).text
    jerseyno = pagesoup.find("div",{"class":"number"}).text
    positionget = pagesoup.find_all("div",{"class":"info"})

    global position
    position = 'null'
    for div in positionget:
        x = div.text
        x = ''.join(x.split())
        if x == 'Goalkeeper':
            position = 'Goalkeeper'
        if x == 'Defender':
            position = 'Defender'
        if x == 'Midfielder':
            position = 'Midfielder'
        if x == 'Forward':
            position = 'Forward'

        print('Position: ' + position)

    team = pagesoup.find("div",{"class":"info"}).text

    if position == 'Defender':
        # determine sheet
        sheet = playerdb.get_sheet_by_name('Defender')

        # determine number of rows existing, write to that row number plus one
        global maxrow
        maxrow = sheet.max_row
        print('Number of Rows ' + str(maxrow))


        #general stats
        appearances = pagesoup.find("span",{"class":"allStatContainer statappearances"}).text
        wins = pagesoup.find("span",{"class":"allStatContainer statwins"}).text
        losses = pagesoup.find("span",{"class":"allStatContainer statlosses"}).text

        #defence stats
        cleansheet = pagesoup.find("span",{"class":"allStatContainer statclean_sheet"}).text
        goalconceded = pagesoup.find("span",{"class":"allStatContainer statgoals_conceded"}).text
        tackles = pagesoup.find("span",{"class":"allStatContainer stattotal_tackle"}).text
        tacklessuccess = pagesoup.find("span",{"class":"allStatContainer stattackle_success"}).text
        lastmantackle = pagesoup.find("span",{"class":"allStatContainer statlast_man_tackle"}).text
        blocks = pagesoup.find("span",{"class":"allStatContainer statblocked_scoring_att"}).text
        interceptions = pagesoup.find("span",{"class":"allStatContainer statinterception"}).text
        clearances = pagesoup.find("span",{"class":"allStatContainer stattotal_clearance"}).text
        headedclearance = pagesoup.find("span",{"class":"allStatContainer stateffective_head_clearance"}).text
        clearanceoffline = pagesoup.find("span",{"class":"allStatContainer statclearance_off_line"}).text
        recovery = pagesoup.find("span",{"class":"allStatContainer statball_recovery"}).text
        duelswon = pagesoup.find("span",{"class":"allStatContainer statduel_won"}).text
        duelslost = pagesoup.find("span",{"class":"allStatContainer statduel_lost"}).text
        fiftyfiftywon = pagesoup.find("span",{"class":"allStatContainer statwon_contest"}).text
        aerialwon = pagesoup.find("span",{"class":"allStatContainer stataerial_won"}).text
        aeriallost = pagesoup.find("span",{"class":"allStatContainer stataerial_lost"}).text
        owngoals = pagesoup.find("span",{"class":"allStatContainer statown_goals"}).text
        errortogoal = pagesoup.find("span",{"class":"allStatContainer staterror_lead_to_goal"}).text


        #discipline stats
        yellowcard = pagesoup.find("span",{"class":"allStatContainer statyellow_card"}).text
        redcard = pagesoup.find("span",{"class":"allStatContainer statred_card"}).text
        fouls = pagesoup.find("span",{"class":"allStatContainer statfouls"}).text
        offsides = pagesoup.find("span",{"class":"allStatContainer stattotal_offside"}).text

        #teamplay stats
        assists = pagesoup.find("span",{"class":"allStatContainer statgoal_assist"}).text
        passes = pagesoup.find("span",{"class":"allStatContainer stattotal_pass"}).text
        passespergame = pagesoup.find("span",{"class":"allStatContainer stattotal_pass_per_game"}).text
        bigchancecreated = pagesoup.find("span", {"class": "allStatContainer statbig_chance_created"}).text
        crosses = pagesoup.find("span", {"class": "allStatContainer stattotal_cross"}).text
        crossacc = pagesoup.find("span", {"class": "allStatContainer statcross_accuracy"}).text
        throughballs = pagesoup.find("span", {"class": "allStatContainer stattotal_through_ball"}).text
        acclongballs = pagesoup.find("span", {"class": "allStatContainer stataccurate_long_balls"}).text

        #attack stats
        goals = pagesoup.find("span",{"class":"allStatContainer statgoals"}).text
        headgoals = pagesoup.find("span",{"class":"allStatContainer statatt_hd_goal"}).text
        goalsright = pagesoup.find("span",{"class":"allStatContainer statatt_rf_goal"}).text
        goalsleft = pagesoup.find("span",{"class":"allStatContainer statatt_lf_goal"}).text
        woodwork = pagesoup.find("span",{"class":"allStatContainer stathit_woodwork"}).text

        #using lists to trim text and for writing to CSV later
        playerdetails = [name, jerseyno, position, team]
        generallist = [appearances, wins, losses]
        defendinglist = [cleansheet,goalconceded,tackles,tacklessuccess,lastmantackle,blocks,interceptions,clearances,headedclearance,clearanceoffline,recovery,duelswon,duelslost,fiftyfiftywon,aerialwon, aeriallost, owngoals, errortogoal]
        disciplinelist = [yellowcard,redcard,fouls, offsides]
        teamplaylist = [assists,passes,passespergame,bigchancecreated,crosses,crossacc, throughballs, acclongballs]
        attacklist = [goals,headgoals,goalsright,goalsleft,woodwork]

        print(len(defendinglist))
        print(len(teamplaylist))
        print(len(attacklist))

        for x,y in zip(playerdetails, range(0,len(playerdetails))):
            if x != name:
                x = ''.join(x.split())
                x = x.replace(',', '')
                print(x)
                playerdetails[y] = x

            if x == name:
                print('Player: ' + name + ' ' + position)
                playerdetails[y] = x

        for x,y in zip(generallist, range(0,len(generallist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            generallist[y] = x

        for x,y in zip(defendinglist, range(0,len(defendinglist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            defendinglist[y] = x

        for x,y in zip(attacklist, range(0,len(attacklist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            attacklist[y] = x

        for x,y in zip(disciplinelist, range(0,len(disciplinelist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            disciplinelist[y] = x

        for x,y in zip(teamplaylist, range(0,len(teamplaylist))):
            x = ''.join(x.split())
            x = x.replace(',','')
            print(x)
            teamplaylist[y] = x

        print(playerdetails)
        print(generallist)
        print(defendinglist)
        print(attacklist)
        print(disciplinelist)
        print(teamplaylist)

        #to remove all whitespace
        # saves = ''.join(saves.split())

        #write to the csv file
        for x,y in zip(playerdetails, range(1,len(playerdetails)+1)):
            sheet.cell(row = maxrow +1, column = y).value = x

        for x,y in zip(generallist, range(1,len(generallist)+1)):
            sheet.cell(row = maxrow +1, column = y + 4).value = x

        for x,y in zip(defendinglist, range(1,len(defendinglist)+1)):
            sheet.cell(row = maxrow +1, column = y + 7).value = x

        for x,y in zip(attacklist, range(1,len(attacklist)+1)):
            sheet.cell(row = maxrow +1, column = y + 25).value = x

        for x,y in zip(disciplinelist, range(1,len(disciplinelist)+1)):
            sheet.cell(row = maxrow +1, column = y + 30).value = x

        for x,y in zip(teamplaylist, range(1,len(teamplaylist)+1)):
            sheet.cell(row = maxrow +1, column = y + 34).value = x


        playerdb.save('EPL 1617.xlsx')

    if position == 'Goalkeeper':
        # determine sheet
        sheet = playerdb.get_sheet_by_name('Goalkeeper')

        # determine number of rows existing, write to that row number plus one
        maxrow = sheet.max_row
        print('Number of Rows ' + str(maxrow))

        # general stats
        appearances = pagesoup.find("span", {"class": "allStatContainer statappearances"}).text
        wins = pagesoup.find("span", {"class": "allStatContainer statwins"}).text
        losses = pagesoup.find("span", {"class": "allStatContainer statlosses"}).text

        # goalkeeping stats
        saves = pagesoup.find("span", {"class": "allStatContainer statsaves"}).text
        penaltysaves = pagesoup.find("span", {"class": "allStatContainer statpenalty_save"}).text
        punches = pagesoup.find("span", {"class": "allStatContainer statpunches"}).text
        highclaim = pagesoup.find("span", {"class": "allStatContainer statgood_high_claim"}).text
        catches = pagesoup.find("span", {"class": "allStatContainer statcatches"}).text
        sweepclearance = pagesoup.find("span", {"class": "allStatContainer stattotal_keeper_sweeper"}).text
        throws = pagesoup.find("span", {"class": "allStatContainer statkeeper_throws"}).text
        goalkicks = pagesoup.find("span", {"class": "allStatContainer statgoal_kicks"}).text

        # defence stats
        cleansheet = pagesoup.find("span", {"class": "allStatContainer statclean_sheet"}).text
        goalconceded = pagesoup.find("span", {"class": "allStatContainer statgoals_conceded"}).text
        errortogoal = pagesoup.find("span", {"class": "allStatContainer staterror_lead_to_goal"}).text
        owngoal = pagesoup.find("span", {"class": "allStatContainer statown_goals"}).text

        # discipline stats
        yellowcard = pagesoup.find("span", {"class": "allStatContainer statyellow_card"}).text
        redcard = pagesoup.find("span", {"class": "allStatContainer statred_card"}).text
        fouls = pagesoup.find("span", {"class": "allStatContainer statfouls"}).text

        # teamplay stats
        goals = pagesoup.find("span", {"class": "allStatContainer statgoals"}).text
        assists = pagesoup.find("span", {"class": "allStatContainer statgoal_assist"}).text
        passes = pagesoup.find("span", {"class": "allStatContainer stattotal_pass"}).text
        passespergame = pagesoup.find("span", {"class": "allStatContainer stattotal_pass_per_game"}).text
        longballs = pagesoup.find("span", {"class": "allStatContainer stataccurate_long_balls"}).text

        # using lists to trim text and for writing to CSV later
        playerdetails = [name, jerseyno, position, team]
        generallist = [appearances, wins, losses]
        goalkeepinglist = [saves, penaltysaves, punches, highclaim, catches, sweepclearance, throws, goalkicks]
        defencelist = [cleansheet, goalconceded, errortogoal, owngoal]
        disciplinelist = [yellowcard, redcard, fouls]
        teamplaylist = [goals, assists, passes, passespergame, longballs]

        for x, y in zip(playerdetails, range(0, len(playerdetails))):
            if x != name:
                x = ''.join(x.split())
                x = x.replace(',', '')
                print(x)
                playerdetails[y] = x

            if x == name:
                print('Player: ' + name + ' ' + position)
                playerdetails[y] = x

        for x, y in zip(generallist, range(0, len(generallist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            generallist[y] = x

        for x, y in zip(goalkeepinglist, range(0, len(goalkeepinglist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            goalkeepinglist[y] = x

        for x, y in zip(defencelist, range(0, len(defencelist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            defencelist[y] = x

        for x, y in zip(disciplinelist, range(0, len(disciplinelist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            disciplinelist[y] = x

        for x, y in zip(teamplaylist, range(0, len(teamplaylist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            teamplaylist[y] = x

        print(playerdetails)
        print(generallist)
        print(goalkeepinglist)
        print(defencelist)
        print(disciplinelist)
        print(teamplaylist)

        # to remove all whitespace
        # saves = ''.join(saves.split())

        # write to the csv file
        for x, y in zip(playerdetails, range(1, len(playerdetails) + 1)):
            sheet.cell(row=maxrow + 1, column=y).value = x

        for x, y in zip(generallist, range(1, len(generallist) + 1)):
            sheet.cell(row=maxrow + 1, column=y + 4).value = x

        for x, y in zip(goalkeepinglist, range(1, len(goalkeepinglist) + 1)):
            sheet.cell(row=maxrow + 1, column=y + 7).value = x

        for x, y in zip(defencelist, range(1, len(defencelist) + 1)):
            sheet.cell(row=maxrow + 1, column=y + 15).value = x

        for x, y in zip(disciplinelist, range(1, len(disciplinelist) + 1)):
            sheet.cell(row=maxrow + 1, column=y + 19).value = x

        for x, y in zip(teamplaylist, range(1, len(teamplaylist) + 1)):
            sheet.cell(row=maxrow + 1, column=y + 22).value = x

        playerdb.save('EPL 1617.xlsx')

    if position == 'Midfielder':
        # determine sheet
        sheet = playerdb.get_sheet_by_name('Midfielder')

        # determine number of rows existing, write to that row number plus one
        maxrow = sheet.max_row
        print('Number of Rows: ' + str(maxrow))


        #general stats
        appearances = pagesoup.find("span",{"class":"allStatContainer statappearances"}).text
        wins = pagesoup.find("span",{"class":"allStatContainer statwins"}).text
        losses = pagesoup.find("span",{"class":"allStatContainer statlosses"}).text

        #defence stats
        tackles = pagesoup.find("span",{"class":"allStatContainer stattotal_tackle"}).text
        tacklessuccess = pagesoup.find("span",{"class":"allStatContainer stattackle_success"}).text
        blocks = pagesoup.find("span",{"class":"allStatContainer statblocked_scoring_att"}).text
        interceptions = pagesoup.find("span",{"class":"allStatContainer statinterception"}).text
        clearances = pagesoup.find("span",{"class":"allStatContainer stattotal_clearance"}).text
        headedclearance = pagesoup.find("span",{"class":"allStatContainer stateffective_head_clearance"}).text
        recovery = pagesoup.find("span",{"class":"allStatContainer statball_recovery"}).text
        duelswon = pagesoup.find("span",{"class":"allStatContainer statduel_won"}).text
        duelslost = pagesoup.find("span",{"class":"allStatContainer statduel_lost"}).text
        fiftyfiftywon = pagesoup.find("span",{"class":"allStatContainer statwon_contest"}).text
        aerialwon = pagesoup.find("span",{"class":"allStatContainer stataerial_won"}).text
        aeriallost = pagesoup.find("span",{"class":"allStatContainer stataerial_lost"}).text
        errortogoal = pagesoup.find("span",{"class":"allStatContainer staterror_lead_to_goal"}).text


        #discipline stats
        yellowcard = pagesoup.find("span",{"class":"allStatContainer statyellow_card"}).text
        redcard = pagesoup.find("span",{"class":"allStatContainer statred_card"}).text
        fouls = pagesoup.find("span",{"class":"allStatContainer statfouls"}).text
        offsides = pagesoup.find("span",{"class":"allStatContainer stattotal_offside"}).text

        #teamplay stats
        assists = pagesoup.find("span",{"class":"allStatContainer statgoal_assist"}).text
        passes = pagesoup.find("span",{"class":"allStatContainer stattotal_pass"}).text
        passespergame = pagesoup.find("span",{"class":"allStatContainer stattotal_pass_per_game"}).text
        bigchancecreated = pagesoup.find("span", {"class": "allStatContainer statbig_chance_created"}).text
        crosses = pagesoup.find("span", {"class": "allStatContainer stattotal_cross"}).text
        crossacc = pagesoup.find("span", {"class": "allStatContainer statcross_accuracy"}).text
        throughballs = pagesoup.find("span", {"class": "allStatContainer stattotal_through_ball"}).text
        acclongballs = pagesoup.find("span", {"class": "allStatContainer stataccurate_long_balls"}).text

        #attack stats
        goals = pagesoup.find("span",{"class":"allStatContainer statgoals"}).text
        goalspg = pagesoup.find("span",{"class":"allStatContainer statgoals_per_game"}).text
        headgoals = pagesoup.find("span",{"class":"allStatContainer statatt_hd_goal"}).text
        goalsright = pagesoup.find("span",{"class":"allStatContainer statatt_rf_goal"}).text
        goalsleft = pagesoup.find("span",{"class":"allStatContainer statatt_lf_goal"}).text
        penaltyscored = pagesoup.find("span",{"class":"allStatContainer statatt_pen_goal"}).text
        freekickscored = pagesoup.find("span",{"class":"allStatContainer statatt_freekick_goal"}).text
        shots = pagesoup.find("span",{"class":"allStatContainer stattotal_scoring_att"}).text
        shotstarget = pagesoup.find("span",{"class":"allStatContainer statontarget_scoring_att"}).text
        shotacc = pagesoup.find("span",{"class":"allStatContainer statshot_accuracy"}).text
        woodwork = pagesoup.find("span",{"class":"allStatContainer stathit_woodwork"}).text
        bigchancemiss = pagesoup.find("span",{"class":"allStatContainer statbig_chance_missed"}).text

        #using lists to trim text and for writing to CSV later
        playerdetails = [name, jerseyno, position, team]
        generallist = [appearances, wins, losses]
        defendinglist = [tackles,tacklessuccess,blocks,interceptions,clearances,headedclearance,recovery,duelswon,duelslost,fiftyfiftywon,aerialwon,aeriallost,errortogoal]
        disciplinelist = [yellowcard,redcard,fouls, offsides]
        teamplaylist = [assists,passes,passespergame,bigchancecreated,crosses,crossacc, throughballs, acclongballs]
        attacklist = [goals,goalspg,headgoals,goalsright,goalsleft,penaltyscored,freekickscored,shots,shotstarget,shotacc,woodwork,bigchancemiss]

        for x,y in zip(playerdetails, range(0,len(playerdetails))):
            if x != name:
                x = ''.join(x.split())
                x = x.replace(',', '')
                print(x)
                playerdetails[y] = x

            if x == name:
                print('Player: ' + name + ' ' + position)
                playerdetails[y] = x

        for x,y in zip(generallist, range(0,len(generallist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            generallist[y] = x

        for x,y in zip(defendinglist, range(0,len(defendinglist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            defendinglist[y] = x

        for x,y in zip(attacklist, range(0,len(attacklist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            attacklist[y] = x

        for x,y in zip(disciplinelist, range(0,len(disciplinelist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            disciplinelist[y] = x

        for x,y in zip(teamplaylist, range(0,len(teamplaylist))):
            x = ''.join(x.split())
            x = x.replace(',','')
            print(x)
            teamplaylist[y] = x

        print(playerdetails)
        print(generallist)
        print(defendinglist)
        print(attacklist)
        print(disciplinelist)
        print(teamplaylist)

        #to remove all whitespace
        # saves = ''.join(saves.split())

        #write to the csv file
        for x,y in zip(playerdetails, range(1,len(playerdetails)+1)):
            sheet.cell(row = maxrow +1, column = y).value = x

        for x,y in zip(generallist, range(1,len(generallist)+1)):
            sheet.cell(row = maxrow +1, column = y + 4).value = x

        for x,y in zip(defendinglist, range(1,len(defendinglist)+1)):
            sheet.cell(row = maxrow +1, column = y + 7).value = x

        for x,y in zip(disciplinelist, range(1,len(disciplinelist)+1)):
            sheet.cell(row = maxrow +1, column = y + 20).value = x

        for x,y in zip(attacklist, range(1,len(attacklist)+1)):
            sheet.cell(row = maxrow +1, column = y + 24).value = x

        for x,y in zip(teamplaylist, range(1,len(teamplaylist)+1)):
            sheet.cell(row = maxrow +1, column = y + 36).value = x


        playerdb.save('EPL 1617.xlsx')


    if position == 'Forward':
        # determine sheet
        sheet = playerdb.get_sheet_by_name('Forward')

        # determine number of rows existing, write to that row number plus one
        maxrow = sheet.max_row
        print('Number of Rows: ' + str(maxrow))


        #general stats
        appearances = pagesoup.find("span",{"class":"allStatContainer statappearances"}).text
        wins = pagesoup.find("span",{"class":"allStatContainer statwins"}).text
        losses = pagesoup.find("span",{"class":"allStatContainer statlosses"}).text

        #defence stats 5
        tackles = pagesoup.find("span",{"class":"allStatContainer stattotal_tackle"}).text
        blocks = pagesoup.find("span",{"class":"allStatContainer statblocked_scoring_att"}).text
        interceptions = pagesoup.find("span",{"class":"allStatContainer statinterception"}).text
        clearances = pagesoup.find("span",{"class":"allStatContainer stattotal_clearance"}).text
        headedclearance = pagesoup.find("span",{"class":"allStatContainer stateffective_head_clearance"}).text

        #discipline stats 4
        yellowcard = pagesoup.find("span",{"class":"allStatContainer statyellow_card"}).text
        redcard = pagesoup.find("span",{"class":"allStatContainer statred_card"}).text
        fouls = pagesoup.find("span",{"class":"allStatContainer statfouls"}).text
        offsides = pagesoup.find("span",{"class":"allStatContainer stattotal_offside"}).text

        #teamplay stats 5
        assists = pagesoup.find("span",{"class":"allStatContainer statgoal_assist"}).text
        passes = pagesoup.find("span",{"class":"allStatContainer stattotal_pass"}).text
        passespergame = pagesoup.find("span",{"class":"allStatContainer stattotal_pass_per_game"}).text
        bigchancecreated = pagesoup.find("span", {"class": "allStatContainer statbig_chance_created"}).text
        crosses = pagesoup.find("span", {"class": "allStatContainer stattotal_cross"}).text

        #attack stats 12
        goals = pagesoup.find("span",{"class":"allStatContainer statgoals"}).text
        goalspg = pagesoup.find("span",{"class":"allStatContainer statgoals_per_game"}).text
        headgoals = pagesoup.find("span",{"class":"allStatContainer statatt_hd_goal"}).text
        goalsright = pagesoup.find("span",{"class":"allStatContainer statatt_rf_goal"}).text
        goalsleft = pagesoup.find("span",{"class":"allStatContainer statatt_lf_goal"}).text
        penaltyscored = pagesoup.find("span",{"class":"allStatContainer statatt_pen_goal"}).text
        freekickscored = pagesoup.find("span",{"class":"allStatContainer statatt_freekick_goal"}).text
        shots = pagesoup.find("span",{"class":"allStatContainer stattotal_scoring_att"}).text
        shotstarget = pagesoup.find("span",{"class":"allStatContainer statontarget_scoring_att"}).text
        shotacc = pagesoup.find("span",{"class":"allStatContainer statshot_accuracy"}).text
        woodwork = pagesoup.find("span",{"class":"allStatContainer stathit_woodwork"}).text
        bigchancemiss = pagesoup.find("span",{"class":"allStatContainer statbig_chance_missed"}).text

        #using lists to trim text and for writing to CSV later
        playerdetails = [name, jerseyno, position, team]
        generallist = [appearances, wins, losses]
        defendinglist = [tackles,blocks,interceptions,clearances,headedclearance]
        disciplinelist = [yellowcard,redcard,fouls, offsides]
        teamplaylist = [assists,passes,passespergame,bigchancecreated,crosses]
        attacklist = [goals,goalspg,headgoals,goalsright,goalsleft,penaltyscored,freekickscored,shots,shotstarget,shotacc,woodwork,bigchancemiss]

        print(len(defendinglist))
        print(len(attacklist))
        print(len(disciplinelist))
        print(len(teamplaylist))

        for x,y in zip(playerdetails, range(0,len(playerdetails))):
            if x != name:
                x = ''.join(x.split())
                x = x.replace(',', '')
                print(x)
                playerdetails[y] = x

            if x == name:
                print('Player: ' + name + ' ' + position)
                playerdetails[y] = x

        for x,y in zip(generallist, range(0,len(generallist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            generallist[y] = x

        for x,y in zip(defendinglist, range(0,len(defendinglist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            defendinglist[y] = x

        for x,y in zip(attacklist, range(0,len(attacklist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            attacklist[y] = x

        for x,y in zip(disciplinelist, range(0,len(disciplinelist))):
            x = ''.join(x.split())
            x = x.replace(',', '')
            print(x)
            disciplinelist[y] = x

        for x,y in zip(teamplaylist, range(0,len(teamplaylist))):
            x = ''.join(x.split())
            x = x.replace(',','')
            print(x)
            teamplaylist[y] = x

        print(playerdetails)
        print(generallist)
        print(defendinglist)
        print(attacklist)
        print(disciplinelist)
        print(teamplaylist)

        #to remove all whitespace
        # saves = ''.join(saves.split())

        #write to the csv file
        for x,y in zip(playerdetails, range(1,len(playerdetails)+1)):
            sheet.cell(row = maxrow +1, column = y).value = x

        for x,y in zip(generallist, range(1,len(generallist)+1)):
            sheet.cell(row = maxrow +1, column = y + 4).value = x

        for x,y in zip(defendinglist, range(1,len(defendinglist)+1)):
            sheet.cell(row = maxrow +1, column = y + 7).value = x

        for x,y in zip(disciplinelist, range(1,len(disciplinelist)+1)):
            sheet.cell(row = maxrow +1, column = y + 12).value = x

        for x,y in zip(attacklist, range(1,len(attacklist)+1)):
            sheet.cell(row = maxrow +1, column = y + 16).value = x

        for x,y in zip(teamplaylist, range(1,len(teamplaylist)+1)):
            sheet.cell(row = maxrow +1, column = y + 28).value = x

        print('////////////////////////////////////////////////')
    playerdb.save('EPL 1617.xlsx')
    browser.close()

def startscrape2():
    playerdb = load_workbook('testbook.xlsx')
    sheet1 = playerdb.get_sheet_by_name('MatchData')

    # determine number of rows existing, write to that row number plus one
    global maxrow
    maxrow = sheet1.max_row
    print('Number of Rows: ' + str(maxrow))

    #url to be scraped
    myurl = urlentry.get()

    uclient = ureq(myurl)

    pagehtml = uclient.read()

    uclient.close()

    #parsing as html
    pagesoup = soup(pagehtml, "html.parser")

    left = pagesoup.find_all("div", {"class": "container left"})
    left = left.find_all("table", {"class": "playerstats lineups table"})

    print(left)

    homegoals = pagesoup.find_all("td", {"class": "player player-a"})
    awaygoals = pagesoup.find_all("td", {"class": "player player-b"})

    coachlist = []
    coaches = pagesoup.find_all("table", {"class": "playerstats lineups table"})
    for div in coaches:
        x = div.text
        x = ''.join(x.split())
        sep = 'Coach:'
        x = x.split(sep, 1)[1]
        print('Home Coach: ' + x)
        coachlist.append(x)

    print(coachlist)

    teams = pagesoup.find_all("h3", {"class": "thick"})
    teamlist = []

    for div in teams:
        x = div.text
        x = ''.join(x.split())
        teamlist.append(x)

    hometeam = teamlist[0]
    awayteam = teamlist[2]

    print('Home Team: ' + hometeam)
    print('Away Team: ' + awayteam)

    hometeamgoals = []
    awayteamgoals = []
    for div in homegoals:
        x = div.text
        x = re.sub("[^0-9]", "", x)
        if len(x) > 2:
            if x[0] == '4':
                x = int(45) + int(x[2])
                x = str(x)
            if x[0] == '9':
                x = int(90) + int(x[2])
                x = str(x)
        if x != '':
            hometeamgoals.append(x)

    for div in awaygoals:
        x = div.text
        x = re.sub("[^0-9]", "", x)
        if x != '':
            awayteamgoals.append(x)

    print(hometeamgoals)
    print(awayteamgoals)

    #this is to determine match outcome
    if len(hometeamgoals) > len(awayteamgoals):
        print(hometeam + ' beat ' + awayteam)
        matchoutcome = 1
    if len(hometeamgoals) < len(awayteamgoals):
        print(awayteam + ' beat ' + hometeam)
        matchoutcome = 2
    if len(hometeamgoals) == len(awayteamgoals):
        print(hometeam + ' drew with ' + awayteam)
        matchoutcome = 3

    #calculate goal difference
    goaldifference = int(len(hometeamgoals)) - int(len(awayteamgoals))


    #write to csv
    sheet1.cell(row=maxrow + 1, column=2).value = hometeam
    sheet1.cell(row=maxrow + 1, column=4).value = awayteam
    sheet1.cell(row=maxrow + 1, column=6).value = len(hometeamgoals)
    sheet1.cell(row=maxrow + 1, column=7).value = len(awayteamgoals)
    if matchoutcome == '1':
        sheet1.cell(row=maxrow + 1, column=8).value = '1'
    if matchoutcome == '2':
        sheet1.cell(row=maxrow + 1, column=8).value = '2'
    if matchoutcome == '3':
        sheet1.cell(row=maxrow + 1, column=8).value = '3'

    sheet1.cell(row=maxrow + 1, column=9).value = goaldifference

    for x,y in zip(hometeamgoals, range(1,len(hometeamgoals)+1)):
        sheet1.cell(row=maxrow + 1, column = 9+y).value = x

    for x,y in zip(awayteamgoals, range(1,len(awayteamgoals)+1)):
        sheet1.cell(row=maxrow + 1, column = 19+y).value = x

    sheet1.cell(row=maxrow + 1, column=30).value = coachlist[0]
    sheet1.cell(row=maxrow + 1, column=31).value = coachlist[1]


    playerdb.save('EPL 1617.xlsx')


######

urlentry = Entry(urlframe, width = 100)
urlentry.grid(row = 1, column = 1)

urlconfirm = Button(urlframe, text = 'Player Scrape', command = startscrape)
urlconfirm.grid(row = 1, column = 2)

urlconfirm2 = Button(urlframe, text = 'Match Scrape', command = startscrape2)
urlconfirm2.grid(row = 2, column = 2)

test = Button(urlframe, text = 'TEST', command = test)
test.grid(row = 3, column = 2)

appearance = Button(urlframe, text = 'APPEARANCE', command = appearance)
appearance.grid(row = 5, column = 2)

multiscraping = Button(urlframe, text = 'Multi-Scraping', command = multiscrape)
multiscraping.grid(row = 4, column = 2)
##



root.mainloop()